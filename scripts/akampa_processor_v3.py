#!/usr/bin/env python3
"""
akampa_processor_v3.py
──────────────────────────────────────────────────────────────────────
Lee el reporte semanal de Cloudbeds (1 archivo XLSX con 4 pestañas)
y el reporte de WeTravel (opcional), y actualiza akampa-data-v3.js
acumulando el histórico semana a semana.

Uso:
  python3 akampa_processor_v3.py \
    --reporte    Reporte_Sales_Intelligence.xlsx \
    --existing   akampa-data-v3.js \
    --output     akampa-data-v3.js \
    --netlify_site  TU_SITE_ID \
    --netlify_token TU_TOKEN

  # Con WeTravel (opcional):
    --wetravel_lv  Reporting_Payments_WeTravel_LaVentana.xlsx
    --wetravel_yuc Reporting_Payments_WeTravel_Yucatan.xlsx
"""

import argparse, json, sys, re
from collections import defaultdict
from datetime import datetime, date as ddate
from pathlib import Path
from zoneinfo import ZoneInfo

# Zona horaria de operación (México opera permanentemente en UTC-6, sin DST desde 2022)
MX_TZ = ZoneInfo('America/Mexico_City')

def now_mx():
    """Hora actual en CDMX, naive (sin tzinfo) para que strftime no incluya offset."""
    return datetime.now(MX_TZ).replace(tzinfo=None)

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("ERROR: pip install openpyxl")

# ── Catálogo de viajes Bahía Mag (fechas fijas) ───────────────────
TRIP_CATALOG = [
    (1,"Ballena Jorobada","2025-2026","2026-01-04","2026-01-07"),
    (2,"Ballena Jorobada","2025-2026","2026-01-08","2026-01-11"),
    (3,"Ballena Jorobada","2025-2026","2026-01-11","2026-01-14"),
    (4,"Ballena Gris","2025-2026","2026-01-15","2026-01-18"),
    (5,"Ballena Gris","2025-2026","2026-01-18","2026-01-21"),
    (6,"Ballena Gris","2025-2026","2026-01-22","2026-01-25"),
    (7,"Ballena Gris","2025-2026","2026-01-25","2026-01-28"),
    (8,"Ballena Gris","2025-2026","2026-01-30","2026-02-02"),
    (9,"Ballena Gris","2025-2026","2026-02-02","2026-02-05"),
    (10,"Ballena Gris","2025-2026","2026-02-05","2026-02-08"),
    (11,"Ballena Gris","2025-2026","2026-02-08","2026-02-11"),
    (12,"Ballena Gris x Kēntro","2025-2026","2026-02-12","2026-02-15"),
    (13,"Ballena Gris","2025-2026","2026-02-15","2026-02-18"),
    (14,"Ballena Gris","2025-2026","2026-02-19","2026-02-22"),
    (15,"Ballena Gris","2025-2026","2026-02-22","2026-02-25"),
    (16,"Ballena Gris","2025-2026","2026-02-26","2026-03-01"),
    (17,"Ballena Gris","2025-2026","2026-03-01","2026-03-04"),
    (18,"Ballena Gris","2025-2026","2026-03-05","2026-03-08"),
    (19,"Ballena Gris","2025-2026","2026-03-08","2026-03-11"),
    (20,"Mi Compa Chava 3.0","2025-2026","2026-03-13","2026-03-16"),
    (21,"Ballena Gris","2025-2026","2026-03-16","2026-03-19"),
    (22,"Ballena Gris","2025-2026","2026-03-19","2026-03-22"),
    (23,"Ballena Gris","2025-2026","2026-03-22","2026-03-25"),
    (24,"Ballena Gris","2025-2026","2026-03-26","2026-03-29"),
    (25,"Ballena Gris","2025-2026","2026-03-29","2026-04-01"),
    (26,"Ocean Safari","2025-2026","2026-04-02","2026-04-05"),
    (27,"Ocean Safari","2025-2026","2026-04-05","2026-04-08"),
    (28,"Retiro Kalea x La Magia del Caos","2025-2026","2026-04-09","2026-04-12"),
    (29,"Ocean Safari","2025-2026","2026-04-12","2026-04-15"),
    (30,"Pablo Voortus","2025-2026","2026-04-15","2026-04-18"),
    (31,"Marlin & Sardine Run","2026-2027","2026-10-15","2026-10-18"),
    (32,"Marlin & Sardine Run","2026-2027","2026-10-18","2026-10-21"),
    (33,"Marlin & Sardine Run","2026-2027","2026-10-22","2026-10-25"),
    (34,"Marlin & Sardine Run","2026-2027","2026-10-25","2026-10-28"),
    (35,"Marlin & Sardine Run","2026-2027","2026-10-29","2026-11-01"),
    (36,"Marlin & Sardine Run","2026-2027","2026-11-01","2026-11-04"),
    (37,"Marlin & Sardine Run","2026-2027","2026-11-05","2026-11-08"),
    (38,"Marlin & Sardine Run","2026-2027","2026-11-08","2026-11-11"),
    (39,"Marlin & Sardine Run","2026-2027","2026-11-12","2026-11-15"),
    (40,"Marlin & Sardine Run","2026-2027","2026-11-15","2026-11-18"),
    (41,"Marlin & Sardine Run","2026-2027","2026-11-19","2026-11-22"),
    (42,"Marlin & Sardine Run","2026-2027","2026-11-22","2026-11-25"),
    (43,"Marlin & Sardine Run","2026-2027","2026-11-26","2026-11-29"),
    (44,"Marlin & Sardine Run","2026-2027","2026-11-29","2026-12-02"),
    (45,"Marlin & Sardine Run","2026-2027","2026-12-03","2026-12-06"),
    (46,"Marlin & Sardine Run","2026-2027","2026-12-06","2026-12-09"),
    (47,"Marlin & Sardine Run","2026-2027","2026-12-10","2026-12-13"),
    (48,"Ballena Jorobada","2026-2027","2026-12-17","2026-12-20"),
    (49,"Ballena Jorobada","2026-2027","2026-12-23","2026-12-26"),
    (50,"Ballena Jorobada","2026-2027","2026-12-30","2027-01-02"),
    (51,"Ballena Jorobada","2026-2027","2027-01-07","2027-01-10"),
    (52,"Ballena Jorobada","2026-2027","2027-01-14","2027-01-17"),
    (53,"Ballena Gris","2026-2027","2027-01-21","2027-01-24"),
    (54,"Ballena Gris","2026-2027","2027-01-28","2027-01-31"),
    (55,"Ballena Gris","2026-2027","2027-02-04","2027-02-07"),
    (56,"Ballena Gris","2026-2027","2027-02-11","2027-02-14"),
    (57,"Ballena Gris","2026-2027","2027-02-18","2027-02-21"),
    (58,"Ballena Gris","2026-2027","2027-02-25","2027-02-28"),
    (59,"Ballena Gris","2026-2027","2027-03-04","2027-03-07"),
    (60,"Ballena Gris","2026-2027","2027-03-18","2027-03-21"),
    (61,"Ballena Gris","2026-2027","2027-03-25","2027-03-28"),
    (62,"Ballena Gris","2026-2027","2027-04-01","2027-04-04"),
    (63,"Ballena Gris","2026-2027","2027-04-08","2027-04-11"),
    (64,"Ballena Gris","2026-2027","2027-04-15","2027-04-18"),
    (65,"Ballena Gris","2026-2027","2027-04-22","2027-04-25"),
]
BUYOUT_IDS = {28}

# ── Helpers ──────────────────────────────────────────────────────
def xl_rows(wb, sheet):
    return list(wb[sheet].iter_rows(values_only=True))

def find_header(rows, col0):
    return next(i for i, r in enumerate(rows) if r[0] == col0)

def parse_date(v):
    if not v: return None
    try: return datetime.strptime(str(v)[:10], '%Y-%m-%d').date()
    except: return None

def iso_week(d):
    iso = d.isocalendar()
    return f"{iso[0]}-W{str(iso[1]).zfill(2)}"

# ── Parse single XLSX with tabs ──────────────────────────────────
def parse_cloudbeds(path):
    wb = load_workbook(path, read_only=True)
    sheets = wb.sheetnames

    # 1. ReservationBalanceDue — Grand Total / Paid / Balance Due per check-in date
    # Columnas: 0=ResNum(puede ser None), 2=Status, 3=Check-In, 6=GrandTotal, 7=Paid, 8=BalanceDue
    rows = xl_rows(wb, 'ReservationBalanceDue')
    hi   = find_header(rows, 'Reservation Number')
    bal  = defaultdict(lambda: {'cobrado':0,'pend':0,'total':0})
    for r in rows[hi+1:]:
        if not r[3]: continue                           # check-in date es obligatorio
        status = str(r[2] or '').strip()
        if status == 'Cancelled': continue              # excluir cancelaciones
        ci = parse_date(r[3])
        if not ci: continue
        try:
            bal[ci]['total']   += float(r[6] or 0)
            bal[ci]['cobrado'] += float(r[7] or 0)
            bal[ci]['pend']    += float(r[8] or 0)
        except: pass

    # 2. CheckinReview — guest and room counts per check-in date
    rows2 = xl_rows(wb, 'CheckinReview')
    hi2   = find_header(rows2, 'Check-In Date')
    ci_data = defaultdict(lambda: {'guests':0,'rooms':0})
    curr = None
    for r in rows2[hi2+1:]:
        if r[0]: curr = parse_date(r[0])
        if curr and r[9] is not None:
            try:
                ci_data[curr]['guests'] += int(r[9] or 0)
                ci_data[curr]['rooms']  += int(r[11] or 0)
            except: pass

    # 3. TotalRevenuePerGuest — monthly summary
    rows3  = xl_rows(wb, 'TotalRevenuePerGuest')
    hi3    = find_header(rows3, 'Stay Date')
    MN     = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    monthly_new = {}
    for r in rows3[hi3+1:]:
        if not r[0] or not str(r[0])[:4].isdigit(): continue
        if not r[1] or r[1] == 0: continue
        ym = str(r[0])[:7]
        yr, mo = ym.split('-')
        season = '2025-2026' if int(yr)==2026 and int(mo)<=9 else '2026-2027'
        monthly_new.setdefault(season, {})[MN[int(mo)-1]] = {
            'm':   MN[int(mo)-1],
            'g':   int(r[2] or 0),
            'occ': round(float(r[6] or 0), 2),
            'rpg': round(float(r[4])) if r[4] and r[4] != '-' else 0,
            'rev': round(float(r[3] or 0))
        }

    # 4. ReservationsByBookingDate (opcional) — weekly/daily paid amounts by booking date
    booking_weekly, booking_weekly_pend, booking_daily = {}, {}, {}
    if 'ReservationsByBookingDate' in sheets:
        booking_weekly, booking_weekly_pend, booking_daily = \
            parse_booking_date_tab(xl_rows(wb, 'ReservationsByBookingDate'))
        print(f"   ReservationsByBookingDate: "
              f"{len(booking_weekly)} semanas · {len(booking_daily)} días")

    # 5. OccupancyStatistics (opcional) — mejora occ% mensual
    occ_monthly = {}
    if 'OccupancyStatistics' in sheets:
        occ_monthly = parse_occupancy_stats_tab(xl_rows(wb, 'OccupancyStatistics'))
        # Patch occ field in monthly_new with real Occupancy Statistics data
        for season, months in monthly_new.items():
            for m_abbr, mdata in months.items():
                ym = _abbr_to_ym(m_abbr, season)
                if ym and ym in occ_monthly:
                    mdata['occ'] = occ_monthly[ym]
        print(f"   OccupancyStatistics: {len(occ_monthly)} meses")

    wb.close()
    return bal, ci_data, monthly_new, booking_weekly, booking_weekly_pend, booking_daily


def parse_booking_date_tab(rows):
    """
    Parsea el tab ReservationsByBookingDate.
    Columnas: 0=Booking Date, 2=Status, 15=Grand Total, 18=Paid, 19=Balance Due
    Retorna: (weekly_paid, weekly_pend, daily_paid)
    """
    try:
        hi = find_header(rows, 'Booking Date Time - Property')
    except StopIteration:
        return {}, {}, {}

    weekly  = defaultdict(float)
    weekly_pend = defaultdict(float)
    daily   = defaultdict(float)

    for r in rows[hi+1:]:
        if not r[0]: continue
        status = str(r[2] or '').strip()
        if status == 'Cancelled': continue

        booking_date = parse_date(str(r[0])[:10].replace('/', '-'))
        if not booking_date: continue

        try:
            grand_total = float(r[15] or 0)
            paid  = float(r[18] or 0)
            pend  = float(r[19] or 0)
        except (TypeError, ValueError):
            continue

        # Excluir correcciones negativas y filas sin monto
        if grand_total <= 0 and paid <= 0 and pend <= 0:
            continue

        wk = iso_week(booking_date)
        ds = str(booking_date)

        if paid > 0:
            weekly[wk] += paid
            daily[ds]  += paid
        if pend > 0:
            weekly_pend[wk] += pend

    # Redondear a enteros
    return (
        {k: round(v) for k, v in weekly.items()},
        {k: round(v) for k, v in weekly_pend.items()},
        {k: round(v) for k, v in daily.items()},
    )


def parse_occupancy_stats_tab(rows):
    """
    Parsea el tab OccupancyStatistics.
    Columnas: 0=Stay Date, 7=Occupancy%
    Retorna dict {YYYY-MM: avg_occ%}
    """
    try:
        hi = find_header(rows, 'Stay Date')
    except StopIteration:
        return {}

    monthly_occ = defaultdict(list)
    for r in rows[hi+1:]:
        if not r[0]: continue
        d = parse_date(str(r[0])[:10])
        if not d: continue
        occ_val = r[7]
        if occ_val is None or occ_val == '-': continue
        try:
            monthly_occ[str(d)[:7]].append(float(occ_val))
        except (TypeError, ValueError):
            pass

    return {ym: round(sum(v)/len(v), 2) for ym, v in monthly_occ.items() if v}


def _abbr_to_ym(abbr, season):
    """Convierte abreviatura de mes + temporada a YYYY-MM."""
    MN = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    if abbr not in MN: return None
    mo = MN.index(abbr) + 1
    # Inferir año de la temporada
    yr_start, yr_end = season.split('-')
    # Temporada "alta" (oct-dic) → primer año; resto → segundo año
    yr = int(yr_start) if mo >= 10 else int(yr_end)
    return f"{yr}-{str(mo).zfill(2)}"

# ── Merge monthly: new data wins, old data preserved ─────────────
def merge_monthly(existing_monthly, new_monthly):
    """
    new_monthly wins for months it contains.
    Months not in new_monthly are kept from existing.
    """
    result = {}
    all_seasons = set(list(existing_monthly.keys()) + list(new_monthly.keys()))
    for season in all_seasons:
        old = {m['m']: m for m in existing_monthly.get(season, [])}
        new = new_monthly.get(season, {})  # dict keyed by month abbr
        # new wins for its months, old kept for the rest
        merged = {**old, **new}
        result[season] = list(merged.values())
    return result

# ── Build trips: new data wins, historical preserved ─────────────
def _merge_bal(bal, dates):
    """Merge multiple balance-due buckets (from different check-in dates) into one."""
    merged = {'cobrado': 0, 'pend': 0, 'total': 0}
    for d in dates:
        if d in bal:
            merged['cobrado'] += bal[d]['cobrado']
            merged['pend']    += bal[d]['pend']
            merged['total']   += bal[d]['total']
    # Devuelve siempre el dict (aunque sea {0,0,0}) para que branch A se ejecute
    # y use ci_data/prev_for_past en vez de borrar el viaje. getReservations no
    # incluye `paid`, así que pend/cobrado vienen frecuentemente en 0.
    return merged

def build_trips(bal, ci_data, existing_trips):
    """
    For each trip in catalog:
    - If new report has data for any check-in date within the trip window → use it
      (handles late check-ins where guest arrives after the official trip start date)
    - If not → keep existing data from previous JSON
    """
    from datetime import timedelta
    today = ddate.today()
    # Index by start date (more robust than ID when catalog order changes)
    existing_map = {t['start']: t for t in existing_trips}

    # Determine the date range covered by the current Cloudbeds report
    # Only use historical data for trips OUTSIDE this range to avoid
    # carrying over stale/corrupted values for trips the report covers
    report_min = min(bal.keys()) if bal else None
    report_max = max(bal.keys()) if bal else None

    trips = []
    weekly = defaultdict(float)

    for tid, name, season, start_s, end_s in TRIP_CATALOG:
        start = datetime.strptime(start_s, '%Y-%m-%d').date()
        end   = datetime.strptime(end_s,   '%Y-%m-%d').date()
        status = 'past' if end < today else 'next' if start <= today else 'future'

        # Collect all check-in dates that fall within this trip's window
        # Window: [start, end) — any guest who checked in during the trip belongs here
        # This handles late check-ins (e.g., guest arrives day 2 of a 3-day trip)
        trip_dates = [start + timedelta(days=i)
                      for i in range(0, (end - start).days)
                      if start + timedelta(days=i) in bal]

        # Find guest/room data: prefer exact start date, then first match
        ci_date = start if start in ci_data else \
                  next((d for d in trip_dates if d in ci_data), None)
        ci   = ci_data.get(ci_date) if ci_date else None
        fin  = _merge_bal(bal, trip_dates) if trip_dates else None

        # Only fall back to historical if this trip is OUTSIDE the current report's range
        # (prevents stale historical values from overriding zero-cobrado trips in the report)
        in_report_range = report_min and report_max and (report_min <= start <= report_max)
        prev = existing_map.get(start_s, {}) if not in_report_range else {}

        # Preservación de guests/rooms para viajes pasados:
        # Cloudbeds Check-in Review reduce su ventana semana a semana —
        # primero pierde check-ins enteros, luego empieza a traer sólo una
        # fracción de los huéspedes de viajes ya cerrados. Para viajes con
        # status='past' tomamos el MÁXIMO entre el dato nuevo y el JSON
        # previo: una vez que un viaje pasó, los huéspedes registrados
        # sólo pueden mantenerse o crecer (correcciones tardías), nunca
        # bajar.
        prev_for_past = existing_map.get(start_s, {}) if status == 'past' else {}

        if fin:
            # New data available — use it
            if ci:
                rooms  = ci['rooms']
                guests = ci['guests']
            elif prev.get('rooms') or prev.get('guests'):
                rooms  = prev.get('rooms', 0)
                guests = prev.get('guests', 0)
            else:
                # Past trip whose ci_data dropped from the report — preserve previous
                rooms  = prev_for_past.get('rooms', 0)
                guests = prev_for_past.get('guests', 0)
            # Para viajes pasados con histórico: el dato histórico (XLSX migration)
            # es más confiable que el API (getReservations no devuelve `paid`/`grandTotal`
            # y los rooms suelen estar limpiados después del checkout).
            # Sólo usamos guests del API como max (puede haber late check-ins legítimos).
            if status == 'past' and prev_for_past:
                rooms   = prev_for_past.get('rooms', rooms)
                guests  = max(guests, prev_for_past.get('guests', 0))
                cap     = prev_for_past.get('cap', 17 if rooms > 15 else 15)
                occ     = prev_for_past.get('occ', round((rooms/cap)*100, 1) if rooms else 0)
                cobrado = prev_for_past.get('cobrado', 0)
                pend    = prev_for_past.get('pend', 0)
            else:
                cap     = 17 if rooms > 15 else 15
                occ     = round((rooms / cap) * 100, 1) if rooms else 0
                cobrado = round(max(0, fin['cobrado']))
                pend    = round(max(0, fin['pend']))
                # getReservations no devuelve `paid`, así que fin['cobrado'] viene en 0.
                # Si el viaje ya tenía cobrado histórico, preservarlo como piso.
                existing_for_trip = existing_map.get(start_s, {})
                if existing_for_trip.get('cobrado'):
                    cobrado = max(cobrado, existing_for_trip['cobrado'])
            # Total: suma directa de componentes (Grand Total puede ser distorsionado
            # por filas de corrección negativas en Cloudbeds)
            total   = cobrado + pend
        elif prev:
            # No new data — preserve existing
            rooms   = prev.get('rooms', 0)
            guests  = prev.get('guests', 0)
            cap     = prev.get('cap', 15)
            occ     = prev.get('occ', 0)
            cobrado = prev.get('cobrado', 0)
            pend    = prev.get('pend', 0)
            total   = prev.get('total', 0)
        else:
            # Brand new trip with no data yet
            rooms = guests = cobrado = pend = total = 0
            cap = 15; occ = 0

        t = {'id':tid,'name':name,'s':season,'start':start_s,'end':end_s,
             'rooms':rooms,'cap':cap,'occ':occ,'guests':guests,'status':status,
             'cobrado':cobrado,'pend':pend,'total':total}

        if tid in BUYOUT_IDS:
            t.update({'buyout':True,'rooms':15,'cap':15,'occ':100,'guests':30})

        trips.append(t)

        # Weekly accumulation
        if cobrado > 0:
            weekly[iso_week(start)] += cobrado

    return trips, dict(weekly)

# ── Fechas conocidas para viajes sin fechas en el nombre (WeTravel) ──
# Clave: substring del trip name (case-insensitive) → (start YYYY-MM-DD, end YYYY-MM-DD)
WETRAVEL_DATE_OVERRIDES = {
    'osom people': ('2026-10-01', '2026-10-04'),
}

# ── Cutoff temporal de WeTravel ───────────────────────────────────
# Cualquier viaje cuya fecha de FIN sea anterior a esta fecha se excluye
# del dashboard. Sirve para no traer operación legacy (ej. Móbulas 2025).
# Cambiar este valor cuando se quiera correr la ventana de visibilidad.
WETRAVEL_MIN_END_DATE = '2026-01-01'

# ── Stub trips: viajes confirmados que aún no tienen pagos en WeTravel ──
# Se agregan al dashboard con $0 cobrado para visibilidad de pipeline.
# Cuando lleguen pagos en futuras semanas, el processor los integra automáticamente
# (siempre que el `match` haga substring case-insensitive sobre el trip name de WeTravel).
# Estructura: {dest_label: [ {match, name, start, end, cap?} ]}
STUB_TRIPS = {
    'Yucatán': [
        {
            'match': 'daniela mendoza',
            'name':  'Daniela Mendoza - Sahaja',
            'start': '2026-11-13',
            'end':   '2026-11-16',
            'cap':   30,
        },
    ],
}

# ── WeTravel parser ───────────────────────────────────────────────
def parse_wetravel(path, dest_label, existing_trips, keyword=None):
    """
    Parsea un export XLSX de WeTravel Payments Reporting.
    Si se pasa `keyword`, solo incluye viajes cuyo nombre contenga esa cadena
    (útil cuando el archivo contiene múltiples destinos mezclados).
    """
    wb   = load_workbook(path, read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    hi   = next(i for i, r in enumerate(rows) if r[0] == 'Date created (UTC)')
    today = ddate.today()

    MES = {'enero':'01','febrero':'02','marzo':'03','abril':'04','mayo':'05',
           'junio':'06','julio':'07','agosto':'08','septiembre':'09',
           'octubre':'10','noviembre':'11','diciembre':'12'}

    by_trip = defaultdict(list)
    for r in rows[hi+1:]:
        if not r[0]: continue
        # Aceptar Successful y Refunded. Para Refunded computamos el neto
        # (Amount - Refunded) que sigue cobrado. Si el refund es total → omitir.
        if r[4] not in ('Successful', 'Refunded'):
            continue
        gross  = float(r[3] or 0)
        refund = float(r[9] or 0)
        amount = gross - refund
        if amount <= 0:
            continue
        trip_name = str(r[21]) if r[21] else 'Sin nombre'
        # Filtrar por keyword si se especificó (ej. 'La Ventana' o 'Yucatan')
        if keyword and keyword.lower() not in trip_name.lower():
            continue
        pdate = parse_date(str(r[0])[:10].replace('/','-'))
        participants = [p.strip() for p in str(r[20]).split(',') if p.strip()] if r[20] else []
        payment = {
            'date': str(pdate), 'amount': amount,
            'participants': participants
        }
        # Solo expone gross/refund cuando hubo reembolso parcial
        if refund > 0:
            payment['gross']  = gross
            payment['refund'] = refund
        by_trip[trip_name].append(payment)

    def extract_dates(name):
        """
        Extrae (start, end) de nombres como:
          "(19-22 abril 2026)"          → 2026-04-19, 2026-04-22
          "(28 mayo -31 mayo 2026)"     → 2026-05-28, 2026-05-31
          "(4 junio - 7 junio 2026)"    → 2026-06-04, 2026-06-07
        """
        # Patrón B: "D1 MES1 - D2 MES2 YYYY"  (mes puede repetirse)
        m = re.search(
            r'\((\d+)\s+(\w+)\s*[-–]\s*(\d+)\s+(\w+)\s+(\d{4})',
            name, re.IGNORECASE
        )
        if m:
            d1, mes1, d2, mes2, yr = (
                m.group(1), m.group(2).lower(),
                m.group(3), m.group(4).lower(), m.group(5)
            )
            mo1 = MES.get(mes1)
            mo2 = MES.get(mes2)
            if mo1 and mo2:
                return f"{yr}-{mo1}-{d1.zfill(2)}", f"{yr}-{mo2}-{d2.zfill(2)}"

        # Patrón A: "D1-D2 MES YYYY"  (mismo mes)
        m = re.search(
            r'\((\d+)[- ]+(\d+)\s+(\w+)\s+(\d{4})',
            name, re.IGNORECASE
        )
        if m:
            d1, d2, mes, yr = m.group(1), m.group(2), m.group(3).lower(), m.group(4)
            mo = MES.get(mes)
            if mo:
                return f"{yr}-{mo}-{d1.zfill(2)}", f"{yr}-{mo}-{d2.zfill(2)}"

        # Patrón C: "D MES YYYY"  (fecha única)
        m = re.search(r'\((\d+)\s+(\w+)\s+(\d{4})', name, re.IGNORECASE)
        if m:
            d1, mes, yr = m.group(1), m.group(2).lower(), m.group(3)
            mo = MES.get(mes)
            if mo:
                return f"{yr}-{mo}-{d1.zfill(2)}", f"{yr}-{mo}-{d1.zfill(2)}"

        return None, None

    # Build trip list — merge with existing to preserve historical payments
    existing_map = {f"{t['start']}_{t['name']}": t for t in existing_trips}
    cutoff = datetime.strptime(WETRAVEL_MIN_END_DATE, '%Y-%m-%d').date()
    trips = []
    seen_keys = set()
    for i, (trip_name, payments) in enumerate(by_trip.items(), 1):
        start_s, end_s = extract_dates(trip_name)
        if not start_s:
            # Buscar override de fechas por nombre conocido
            name_lower = trip_name.lower()
            override = next(
                (v for k, v in WETRAVEL_DATE_OVERRIDES.items() if k in name_lower),
                None
            )
            if override:
                start_s, end_s = override
            else:
                start_s = payments[0]['date'] if payments else str(today)
                end_s   = start_s
        start  = datetime.strptime(start_s, '%Y-%m-%d').date()
        end    = datetime.strptime(end_s,   '%Y-%m-%d').date()
        # Excluir viajes legacy cuya fecha fin sea anterior al cutoff configurado
        if end < cutoff:
            continue
        status = 'past' if end < today else 'next' if start <= today else 'future'
        clean  = re.sub(r'\s*\(.*?\)\s*$', '', trip_name).strip()  # quita fecha al final
        clean  = re.sub(r'^[^:]+:\s*', '', clean)                   # quita "Destino: "
        clean  = re.sub(r'\s*\|.*$', '', clean).strip()             # quita " | Camp name"
        trips.append({'id':i,'name':clean,'dest':dest_label,'start':start_s,
                      'end':end_s,'cap':30,'status':status,'payments':payments})
        seen_keys.add(f"{start_s}_{clean}")

    # Preservar viajes históricos pasados que existían en el JSON previo y
    # que el export actual de WeTravel ya no trae. WeTravel limita la ventana
    # de su export semana a semana, pero el dashboard debe mantener el
    # histórico para no perder revenue acumulado.
    for prev_t in existing_trips:
        key = f"{prev_t.get('start')}_{prev_t.get('name')}"
        if key in seen_keys:
            continue
        end_s = prev_t.get('end') or prev_t.get('start')
        try:
            end_d = datetime.strptime(end_s, '%Y-%m-%d').date()
        except (TypeError, ValueError):
            continue
        if end_d < cutoff:
            continue
        # Sólo conservar viajes que ya pasaron — los futuros que no estén en el
        # export actual probablemente fueron cancelados/movidos.
        if end_d >= today:
            continue
        kept = dict(prev_t)
        kept['status'] = 'past'
        trips.append(kept)
        seen_keys.add(key)

    # ── Agregar stub trips que no estén ya representados por pagos ────
    for stub in STUB_TRIPS.get(dest_label, []):
        match = stub['match'].lower()
        already = any(match in t['name'].lower() for t in trips)
        if already:
            continue
        s_start = stub['start']
        s_end   = stub['end']
        s_start_d = datetime.strptime(s_start, '%Y-%m-%d').date()
        s_end_d   = datetime.strptime(s_end,   '%Y-%m-%d').date()
        s_status  = 'past' if s_end_d < today else 'next' if s_start_d <= today else 'future'
        trips.append({
            'id': len(trips) + 1,
            'name': stub['name'],
            'dest': dest_label,
            'start': s_start,
            'end':   s_end,
            'cap':   stub.get('cap', 30),
            'status': s_status,
            'payments': [],
        })

    trips.sort(key=lambda t: t['start'])
    # Re-numerar IDs después del sort para mantener consistencia
    for i, t in enumerate(trips, 1):
        t['id'] = i
    return trips

# ── HTML update helpers ───────────────────────────────────────────
def _fmt_bm_trips(trips):
    """Serializa BM_TRIPS como literal JS de una línea por viaje."""
    lines = ['const BM_TRIPS = [']
    for t in trips:
        parts = [
            f'id:{t["id"]}',
            f'name:{json.dumps(t["name"], ensure_ascii=False)}',
            f's:{json.dumps(t["s"])}',
            f'start:{json.dumps(t["start"])}',
            f'end:{json.dumps(t["end"])}',
            f'rooms:{t.get("rooms", 0)}',
            f'cap:{t.get("cap", 15)}',
            f'occ:{t.get("occ", 0)}',
            f'guests:{t.get("guests", 0)}',
            f'status:{json.dumps(t["status"])}',
            f'cobrado:{t.get("cobrado", 0)}',
            f'pend:{t.get("pend", 0)}',
            f'total:{t.get("total", 0)}',
        ]
        if t.get('buyout'):
            parts.append('buyout:true')
        lines.append('  {' + ', '.join(parts) + '},')
    lines.append('];')
    return '\n'.join(lines)


def _fmt_lv_trips(trips, var_name='LV_TRIPS'):
    """Serializa LV_TRIPS / YUC_TRIPS con pagos anidados."""
    if not trips:
        return f'const {var_name} = [];'
    parts = [f'const {var_name} = [']
    for t in trips:
        pmts = []
        for p in t.get('payments', []):
            plist = json.dumps(p['participants'], ensure_ascii=False)
            fields = [
                f'date:{json.dumps(p["date"])}',
                f'amount:{p["amount"]}',
                f'participants:{plist}',
            ]
            if 'refund' in p:
                fields.insert(2, f'gross:{p["gross"]}')
                fields.insert(3, f'refund:{p["refund"]}')
            pmts.append('      {' + ','.join(fields) + '}')
        pmts_str = ',\n'.join(pmts)
        parts.append(
            f'  {{\n'
            f'    id:{t["id"]}, name:{json.dumps(t["name"], ensure_ascii=False)}, '
            f'dest:{json.dumps(t.get("dest", ""))},\n'
            f'    start:{json.dumps(t["start"])}, end:{json.dumps(t["end"])}, '
            f'cap:{t.get("cap", 30)}, status:{json.dumps(t["status"])},\n'
            f'    payments:[\n{pmts_str}\n    ]\n'
            f'  }},'
        )
    parts.append('];')
    return '\n'.join(parts)


def _fmt_monthly(monthly):
    """Serializa BM_MONTHLY. Entrada: dict {season: [meses...]}."""
    # Asegurar que los valores sean listas (no dicts)
    clean = {}
    for season, v in monthly.items():
        if isinstance(v, dict):
            clean[season] = list(v.values())
        else:
            clean[season] = v
    return 'const BM_MONTHLY = ' + json.dumps(clean, ensure_ascii=False, indent=2) + ';'


def build_html_data_block(today_str, trips, monthly, weekly, weekly_pend, daily,
                           lv_trips, yuc_trips):
    """Genera el bloque completo entre marcadores AKAMPA:DATA_START/END."""
    return (
        '// AKAMPA:DATA_START\n'
        f'const META = {{ kpi: 30000000, today: \'{today_str}\' }};\n'
        '\n'
        '// Bahía Mag trips\n'
        + _fmt_bm_trips(trips) + '\n'
        '\n'
        '// La Ventana trips (WeTravel)\n'
        + _fmt_lv_trips(lv_trips, 'LV_TRIPS') + '\n'
        '\n'
        '// Yucatán trips\n'
        + _fmt_lv_trips(yuc_trips, 'YUC_TRIPS') + '\n'
        '\n'
        f'// Bahía Mag — ventas por semana de BOOKING DATE (Cloudbeds · {today_str})\n'
        f'const BM_WEEKLY = {json.dumps(weekly, ensure_ascii=False)};\n'
        '\n'
        '// Bahía Mag — balance due pendiente por semana de booking\n'
        f'const BM_WEEKLY_PEND = {json.dumps(weekly_pend, ensure_ascii=False)};\n'
        '\n'
        '// Bahía Mag — ventas diarias por BOOKING DATE\n'
        f'const BM_DAILY  = {json.dumps(daily, ensure_ascii=False)};\n'
        '\n'
        + _fmt_monthly(monthly) + '\n'
        '// AKAMPA:DATA_END'
    )


def _fmt_date_es(date_str):
    """Formatea '2026-04-17' → '17 abr 2026'."""
    meses = ['ene','feb','mar','abr','may','jun','jul','ago','sep','oct','nov','dic']
    try:
        d = datetime.strptime(date_str, '%Y-%m-%d')
        return f"{d.day} {meses[d.month-1]} {d.year}"
    except:
        return date_str

def update_html(html_path, data_block, today_str=None, report_max_str=None):
    """Reemplaza la sección AKAMPA:DATA_START…END en index.html
    y actualiza el timestamp visible en el header."""
    content = html_path.read_text(encoding='utf-8')
    if '// AKAMPA:DATA_START' not in content:
        print(f'⚠  No se encontró marcador AKAMPA:DATA_START en {html_path.name} — saltando')
        return False

    # 1. Replace data block (use lambda to avoid \u escape issues in replacement)
    updated = re.sub(
        r'// AKAMPA:DATA_START.*?// AKAMPA:DATA_END',
        lambda m: data_block,
        content,
        flags=re.DOTALL
    )

    # 2. Update visible timestamp in header
    if today_str:
        fecha_str = _fmt_date_es(today_str)
        now_time  = now_mx().strftime('%H:%M')
        if report_max_str and report_max_str != today_str:
            datos_str = _fmt_date_es(report_max_str)
            label = f'Actualizado {fecha_str} · {now_time} · Datos al {datos_str}'
        else:
            label = f'Actualizado {fecha_str} · {now_time}'
        updated = re.sub(
            r'(<span[^>]*id="last-updated"[^>]*>)[^<]*(</span>)',
            rf'\g<1>{label}\g<2>',
            updated
        )

    html_path.write_text(updated, encoding='utf-8')
    print(f'✓ {html_path.name} actualizado')
    return True


# ── Netlify deploy ────────────────────────────────────────────────
def deploy_to_netlify(js_path, site_id, token):
    """Upload both index.html and akampa-data-v3.js via Netlify deploy API."""
    import urllib.request, urllib.error, hashlib
    from pathlib import Path as _P

    # Load both files
    js_content   = _P(js_path).read_bytes()
    js_sha1      = hashlib.sha1(js_content).hexdigest()

    # Find index.html — same folder as the .js file
    html_path = _P(js_path).parent / 'akampa-dashboard-v3.html'
    if not html_path.exists():
        # Try common locations
        for candidate in [
            _P.home() / 'Desktop' / 'akampa-deploy' / 'index.html',
            _P.home() / 'Documents' / 'Cowork' / 'akampa' / 'akampa-dashboard-v3.html',
        ]:
            if candidate.exists():
                html_path = candidate
                break

    if html_path.exists():
        html_content = html_path.read_bytes()
        html_sha1    = hashlib.sha1(html_content).hexdigest()
        files_manifest = {
            '/index.html':         html_sha1,
            '/akampa-data-v3.js':  js_sha1,
        }
        print(f"   Subiendo index.html + akampa-data-v3.js")
    else:
        html_content = None
        files_manifest = {'/akampa-data-v3.js': js_sha1}
        print(f"   ⚠ index.html no encontrado — subiendo solo .js")

    base = 'https://api.netlify.com/api/v1'

    def req(method, url, body=None, raw=None, ct='application/json'):
        data = raw if raw else (json.dumps(body).encode() if body else None)
        r = urllib.request.Request(url, data=data, method=method)
        r.add_header('Authorization', f'Bearer {token}')
        r.add_header('Content-Type',  ct)
        try:
            with urllib.request.urlopen(r, timeout=30) as resp:
                return json.loads(resp.read()), resp.status
        except urllib.error.HTTPError as e:
            return json.loads(e.read().decode() or '{}'), e.code

    # 1. Create deploy with file manifest
    resp, status = req('POST', f'{base}/sites/{site_id}/deploys',
                       body={'files': files_manifest, 'async': False})
    if status not in (200, 201):
        print(f"❌ Netlify error {status}: {resp}")
        return False

    deploy_id = resp['id']
    required  = resp.get('required', [])
    print(f"   Deploy {deploy_id[:12]}... creado ({len(required)} archivos requeridos)")

    # 2. Upload required files
    for sha, content_bytes, filename in [
        (js_sha1,   js_content,   'akampa-data-v3.js'),
        (html_sha1 if html_content else None, html_content, 'index.html'),
    ]:
        if not content_bytes: continue
        if required and sha not in required: continue
        resp2, s2 = req('PUT', f'{base}/deploys/{deploy_id}/files/{filename}',
                        raw=content_bytes, ct='application/octet-stream')
        if s2 not in (200, 201):
            print(f"❌ Upload error {filename} {s2}: {resp2}")
            return False
        print(f"   ✓ {filename} subido")

    print(f"✅ Netlify actualizado — https://akampa-sales.netlify.app")
    return True

# ── Main ─────────────────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser(description='Akampa v3 weekly processor')
    # ── Fuente Cloudbeds: XLSX o API (mutuamente excluyentes) ─────
    src = p.add_mutually_exclusive_group()
    src.add_argument('--reporte',            default=None,  help='XLSX combinado de Cloudbeds')
    src.add_argument('--use_api',            action='store_true', help='Usar Cloudbeds API en lugar de XLSX')
    # Args solo para modo API
    p.add_argument('--api_key',              default=None,  help='Bearer token Cloudbeds (cbat_...)')
    p.add_argument('--desde',               default='2026-01-01')
    p.add_argument('--hasta',               default='2027-12-31')
    p.add_argument('--desde_booking',       default='2025-01-01')
    # ── WeTravel ─────────────────────────────────────────────────
    p.add_argument('--wetravel_lv',          default=None,  help='XLSX WeTravel La Ventana (o combinado)')
    p.add_argument('--wetravel_lv_keyword',  default='La Ventana')
    p.add_argument('--wetravel_yuc',         default=None,  help='XLSX WeTravel Yucatán (puede ser el mismo combinado)')
    p.add_argument('--wetravel_yuc_keyword', default='Yucatan')
    # ── Salida ────────────────────────────────────────────────────
    p.add_argument('--existing',             default=None,  help='akampa-data-v3.js anterior')
    p.add_argument('--output',               default='akampa-data-v3.js')
    p.add_argument('--html',                 default=None,  help='Ruta a index.html (para actualizar datos embebidos)')
    p.add_argument('--netlify_site',         default=None)
    p.add_argument('--netlify_token',        default=None)
    args = p.parse_args()

    # Validar que hay al menos una fuente Cloudbeds
    if not args.use_api and not args.reporte:
        p.error('Se requiere --reporte (XLSX) o --use_api (API)')

    _now_mx   = now_mx()
    today_str = _now_mx.strftime('%Y-%m-%d')
    now_time  = _now_mx.strftime('%H:%M')

    # ── Load existing data ────────────────────────────────────────
    existing = {
        'bahia_mag': {'trips':[],'monthly':{},'weekly':{},'weekly_pend':{},'daily':{}},
        'la_ventana': {'trips':[]},
        'yucatan':    {'trips':[]}
    }
    if args.existing:
        ep = Path(args.existing).expanduser()
        if ep.exists():
            raw = ep.read_text(encoding='utf-8')
            # Quitar comentarios iniciales y el wrapper window.AKAMPA_DATA = ...;
            raw = re.sub(r'.*?window\.AKAMPA_DATA\s*=\s*', '', raw, flags=re.DOTALL).rstrip(';')
            try:
                existing = json.loads(raw)
                print(f"📂 Histórico cargado: {len(existing['bahia_mag']['trips'])} viajes BM")
            except Exception as e:
                print(f"⚠ No se pudo leer histórico: {e} — comenzando desde cero")

    # ── Parse Cloudbeds: API o XLSX ───────────────────────────────
    if args.use_api:
        import importlib.util, os
        _script_dir = Path(__file__).parent
        _api_path   = _script_dir / 'cloudbeds_api.py'
        spec = importlib.util.spec_from_file_location('cloudbeds_api', _api_path)
        cb_mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(cb_mod)

        api_key = args.api_key or os.environ.get('CLOUDBEDS_API_KEY', '')
        if not api_key:
            print('❌ Se necesita --api_key o la variable CLOUDBEDS_API_KEY')
            sys.exit(1)

        print(f"🌐 Fetching datos desde Cloudbeds API ({args.desde} → {args.hasta})...")
        bal, ci_data, monthly_new, new_booking_wk, new_booking_pend, new_booking_daily = \
            cb_mod.fetch_cloudbeds_api(
                api_key       = api_key,
                desde         = args.desde,
                hasta         = args.hasta,
                desde_booking = args.desde_booking,
            )
    else:
        print(f"📂 Procesando reporte Cloudbeds: {args.reporte}")
        bal, ci_data, monthly_new, new_booking_wk, new_booking_pend, new_booking_daily = \
            parse_cloudbeds(args.reporte)

    print(f"   Balance Due: {len(bal)} fechas · Check-in: {len(ci_data)} fechas "
          f"· Monthly: {sum(len(v) for v in monthly_new.values())} meses")

    # Merge monthly (new wins)
    merged_monthly = merge_monthly(existing['bahia_mag'].get('monthly', {}), monthly_new)

    # Build trips (new wins, historical preserved)
    trips, _ = build_trips(bal, ci_data, existing['bahia_mag'].get('trips', []))

    # Booking-date weekly (from ReservationsByBookingDate — replaces check-in based)
    old_bk_wk   = existing['bahia_mag'].get('weekly',       {})
    old_bk_pend = existing['bahia_mag'].get('weekly_pend',  {})
    old_bk_daily= existing['bahia_mag'].get('daily',        {})

    if new_booking_wk or new_booking_pend or new_booking_daily:
        # New report wins for all weeks it covers; old weeks preserved for the rest
        merged_bk_wk    = {**old_bk_wk,    **new_booking_wk}
        merged_bk_pend  = {**old_bk_pend,  **new_booking_pend}
        merged_bk_daily = {**old_bk_daily, **new_booking_daily}
        print(f"   Booking weekly: {len(merged_bk_wk)} semanas · "
              f"Pend: {len(merged_bk_pend)} · Daily: {len(merged_bk_daily)} días")
    else:
        merged_bk_wk    = old_bk_wk
        merged_bk_pend  = old_bk_pend
        merged_bk_daily = old_bk_daily
        print(f"   Booking weekly: conservando {len(merged_bk_wk)} semanas anteriores")

    print(f"   Trips: {len(trips)}")

    # ── WeTravel ─────────────────────────────────────────────────
    if args.wetravel_lv:
        print(f"📂 Procesando WeTravel La Ventana (keyword: '{args.wetravel_lv_keyword}')...")
        lv_trips = parse_wetravel(args.wetravel_lv, 'La Ventana',
                                  existing['la_ventana'].get('trips', []),
                                  keyword=args.wetravel_lv_keyword)
        print(f"   {len(lv_trips)} viajes")
    else:
        lv_trips = existing['la_ventana'].get('trips', [])
        print(f"   La Ventana: conservando {len(lv_trips)} viajes existentes")

    if args.wetravel_yuc:
        print(f"📂 Procesando WeTravel Yucatán (keyword: '{args.wetravel_yuc_keyword}')...")
        yuc_trips = parse_wetravel(args.wetravel_yuc, 'Yucatán',
                                   existing['yucatan'].get('trips', []),
                                   keyword=args.wetravel_yuc_keyword)
        print(f"   {len(yuc_trips)} viajes")
    else:
        yuc_trips = existing['yucatan'].get('trips', [])
        print(f"   Yucatán: conservando {len(yuc_trips)} viajes existentes")

    # ── Build final data object ───────────────────────────────────
    data = {
        'meta': {
            'kpi_anual':    30000000,
            'last_updated': f'{today_str} {now_time}',
            'property':     'Akampa · All Destinations'
        },
        'bahia_mag': {
            'trips':        trips,
            'monthly':      merged_monthly,
            'weekly':       merged_bk_wk,
            'weekly_pend':  merged_bk_pend,
            'daily':        merged_bk_daily,
        },
        'la_ventana': {'trips': lv_trips},
        'yucatan':    {'trips': yuc_trips}
    }

    # ── Write akampa-data-v3.js ───────────────────────────────────
    out = Path(args.output).expanduser()
    # Header comment with generation date
    header = f'// akampa-data-v3.js — generado {today_str}\n'
    header += '// Solo expone window.AKAMPA_DATA (sin conflictos de scope)\n'
    js = header + 'window.AKAMPA_DATA = ' + json.dumps(data, ensure_ascii=False, indent=2) + ';'
    out.write_text(js, encoding='utf-8')
    print(f"\n✅ {out.name} actualizado — {today_str}")

    # ── Update index.html ─────────────────────────────────────────
    if args.html:
        hp = Path(args.html).expanduser()
        if hp.exists():
            block = build_html_data_block(
                today_str, trips, merged_monthly,
                merged_bk_wk, merged_bk_pend, merged_bk_daily,
                lv_trips, yuc_trips
            )
            # Pass report_max so the timestamp shows "Datos al X"
            report_max_str = str(max(bal.keys())) if bal else None
            update_html(hp, block, today_str=today_str, report_max_str=report_max_str)
        else:
            print(f"⚠  --html apunta a un archivo que no existe: {hp}")

    # ── Deploy to Netlify (legacy) ────────────────────────────────
    if args.netlify_site and args.netlify_token:
        print("🚀 Subiendo a Netlify...")
        deploy_to_netlify(str(out), args.netlify_site, args.netlify_token)
    else:
        print("ℹ  Git deploy activo — Netlify no configurado")

if __name__ == '__main__':
    main()
