#!/usr/bin/env python3
"""
combine_cloudbeds.py
────────────────────────────────────────────────────────────────
Combina los reportes individuales de Cloudbeds (descargados por
separado) en un solo XLSX con las pestañas que espera el procesador.

Luego corre akampa_processor_v3.py con todos los datos acumulados.

Uso:
  python3 scripts/combine_cloudbeds.py
"""

import glob, subprocess, sys, unicodedata
from pathlib import Path
from openpyxl import load_workbook, Workbook

REPO = Path(__file__).parent.parent
REPORTES = REPO / 'reportes'

# ── Mapeo: lista de patrones (EN + ES) → nombre de pestaña esperado ─
# Cloudbeds puede exportar los reportes con nombre en inglés o en español
# dependiendo del idioma de la cuenta. Aceptamos ambos.
TAB_MAP = {
    'ReservationBalanceDue':      ['Reservation Balance Due',      'Saldo pendiente de reserva'],
    'CheckinReview':              ['Check-in Review',              'Revisión de entradas'],
    'TotalRevenuePerGuest':       ['Total Revenue Per Guest',      'Ingresos totales por huésped'],
    'ReservationsByBookingDate':  ['Reservations by Booking Date', 'Reservas por fecha de reserva'],
    'OccupancyStatistics':        ['Occupancy Statistics',         'Estadísticas de ocupación'],
}

def _norm(s):
    """Normaliza Unicode a NFC y minúsculas. macOS guarda nombres con acentos
    en NFD (descompuestos: 'o' + U+0301), pero los keywords del script están en
    NFC ('ó'). Sin normalizar, el `in` falla en archivos como 'Revisión...'
    o 'Estadísticas...'."""
    return unicodedata.normalize('NFC', s).lower()

def find_file(keywords):
    """Busca el archivo más reciente en reportes/ cuyo nombre contenga
    cualquiera de los keywords (ej. nombre EN o ES de Cloudbeds).
    Acepta tanto string como lista de strings.
    Selecciona el archivo con la fecha de modificación más reciente para
    evitar problemas cuando los nombres mezclan EN/ES de distintas semanas."""
    if isinstance(keywords, str):
        keywords = [keywords]
    norm_kw = [_norm(k) for k in keywords]
    matches = []
    for f in REPORTES.glob('*.xlsx'):
        if f.name.startswith('~$'):
            continue
        nf = _norm(f.name)
        if any(k in nf for k in norm_kw):
            matches.append(f)
    return max(matches, key=lambda f: f.stat().st_mtime) if matches else None

def copy_sheet(src_path, dest_wb, tab_name):
    """Lee la primera hoja del archivo fuente y la copia al workbook destino."""
    src_wb = load_workbook(src_path, read_only=True, data_only=True)
    src_ws = src_wb.active
    dest_ws = dest_wb.create_sheet(title=tab_name)
    for row in src_ws.iter_rows(values_only=True):
        dest_ws.append(list(row))
    src_wb.close()
    print(f"  ✓ {tab_name} ← {src_path.name}")

def main():
    print("=" * 60)
    print("🔧 Combinando reportes de Cloudbeds...")
    print("=" * 60)

    combined_wb = Workbook()
    combined_wb.remove(combined_wb.active)  # quitar hoja vacía default

    missing = []
    for tab_name, keywords in TAB_MAP.items():
        f = find_file(keywords)
        if f:
            copy_sheet(f, combined_wb, tab_name)
        else:
            kw_str = ' | '.join(keywords)
            print(f"  ⚠ No encontrado: ningún archivo con '{kw_str}' en reportes/")
            missing.append(tab_name)

    if 'ReservationBalanceDue' in missing:
        print("\n❌ El archivo de Reservation Balance Due es obligatorio.")
        sys.exit(1)

    combined_path = REPORTES / 'cloudbeds_combined.xlsx'
    combined_wb.save(combined_path)
    print(f"\n✓ Archivo combinado → {combined_path.name}")

    # ── Buscar WeTravel ───────────────────────────────────────────
    wt_file = find_file('Reporting_Payments_WeTravel')
    if wt_file:
        print(f"✓ WeTravel: {wt_file.name}")
    else:
        print("  ℹ WeTravel no encontrado — se conservan datos anteriores")

    # ── Correr el procesador ──────────────────────────────────────
    print("\n" + "=" * 60)
    print("⚙️  Procesando datos...")
    print("=" * 60)

    existing_js = REPO / 'akampa-data-v3.js'
    html_path   = REPO / 'index.html'
    cmd = [
        sys.executable, str(REPO / 'scripts' / 'akampa_processor_v3.py'),
        '--reporte',  str(combined_path),
        '--existing', str(existing_js),
        '--output',   str(existing_js),
        '--html',     str(html_path),
    ]
    if wt_file:
        cmd += [
            '--wetravel_lv',          str(wt_file),
            '--wetravel_lv_keyword',  'La Ventana',
            '--wetravel_yuc',         str(wt_file),
            '--wetravel_yuc_keyword', 'Yucatan Jungle Camp',
        ]

    result = subprocess.run(cmd, cwd=str(REPO))
    if result.returncode != 0:
        print("❌ Error en el procesador.")
        sys.exit(1)

    print("\n" + "=" * 60)
    print("🚀 Publicando en GitHub...")
    print("=" * 60)
    subprocess.run(
        'git add akampa-data-v3.js akampa-data-v3.json index.html && '
        'git diff --staged --quiet && echo "ℹ Sin cambios nuevos." || '
        '(git commit -m "🔄 Auto-update $(date +%Y-%m-%d)" && git push)',
        shell=True, cwd=str(REPO)
    )

    print("\n✅ Listo → https://sales.akampa.mx")

if __name__ == '__main__':
    main()
